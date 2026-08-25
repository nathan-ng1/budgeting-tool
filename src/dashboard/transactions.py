"""Translation between the Transactions tab's JSON shape and Transaction/
Candidate - see Issues #33 and #35.

Kept out of dashboard.server so the HTTP layer stays a router: what a
Transaction looks like on the wire is a question about the domain, not about
HTTP - mirrors dashboard.recurring.
"""

import base64
import binascii
import csv
import io
import zipfile
from datetime import date, datetime

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

from transaction_log.categories import TYPE_ORDER, Category, assignable_categories_by_type, type_lookup
from transaction_log.entries import Candidate, ExistingRow, Transaction
from transaction_log.writer import resolve_writes

FIELDS = ("date", "amount", "type", "category", "notes")
CSV_HEADER = ("Date", "Amount", "Type", "Category", "Notes")

# How many pre-formatted data rows the Import template offers - generous
# enough for a normal import batch without the workbook growing unwieldy.
IMPORT_TEMPLATE_ROWS = 500

DATE_NUMBER_FORMAT = "YYYY-MM-DD"

_COLUMN_INSTRUCTIONS = (
    ("Date", "The Transaction's date.", "An Excel date, or unambiguous YYYY-MM-DD text."),
    ("Amount", "The Transaction's amount.", "A positive number."),
    ("Type", "Income, Expense, Debt, or Transfer.", "Choose from the dropdown."),
    (
        "Category",
        "The specific budget label for the Transaction.",
        "Choose from the dropdown - see the Type / Category table below for which "
        "Categories belong to which Type.",
    ),
    ("Notes", "A description of the Transaction (e.g. the merchant).", "Free text."),
)


def as_payload(transaction: Transaction) -> dict:
    return {
        "id": transaction.id,
        "date": transaction.date.isoformat(),
        "amount": transaction.amount,
        "type": transaction.type,
        "category": transaction.category,
        "notes": transaction.notes,
    }


def as_csv(transactions: list[Transaction]) -> bytes:
    """The Export panel's `.csv` (Issue #96) - the same 5 columns as the
    Transaction Log, header-only when `transactions` is empty, with no
    internal `id` since that's meaningless outside the Dashboard.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(CSV_HEADER)
    for t in transactions:
        writer.writerow([t.date.isoformat(), t.amount, t.type, t.category, t.notes])
    return buffer.getvalue().encode("utf-8")


def as_import_template(categories: list[Category]) -> bytes:
    """The Import panel's `.xlsx` template (Issue #97) - generated fresh on
    every request from the live Category list, so it never goes stale as
    Categories are added, renamed, or removed via Category Management.
    Upload/preview/commit is a follow-on ticket - this only has to be
    correct and downloadable.
    """
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(CSV_HEADER)

    assignable = assignable_categories_by_type(categories)
    type_names = list(TYPE_ORDER)
    category_names = sorted({name for names in assignable.values() for name in names})

    instructions = workbook.create_sheet("Instructions")
    _write_instructions(instructions, categories)

    # A hidden sheet holding the dropdown source lists - a flat (non-dependent)
    # list per column, per the AC, not a Type->Category cascade. Excel data
    # validation reads a list from a real range rather than an inline string,
    # which openpyxl's DataValidation also caps at 255 characters.
    lists_sheet = workbook.create_sheet("Lists")
    lists_sheet.sheet_state = "hidden"
    for row, type_name in enumerate(type_names, start=1):
        lists_sheet.cell(row=row, column=1, value=type_name)
    for row, category_name in enumerate(category_names, start=1):
        lists_sheet.cell(row=row, column=2, value=category_name)

    last_row = IMPORT_TEMPLATE_ROWS + 1

    type_validation = DataValidation(
        type="list", formula1=f"Lists!$A$1:$A${len(type_names)}", allow_blank=True
    )
    sheet.add_data_validation(type_validation)
    type_validation.add(f"C2:C{last_row}")

    if category_names:
        category_validation = DataValidation(
            type="list", formula1=f"Lists!$B$1:$B${len(category_names)}", allow_blank=True
        )
        sheet.add_data_validation(category_validation)
        category_validation.add(f"D2:D{last_row}")

    for row in range(2, last_row + 1):
        sheet.cell(row=row, column=1).number_format = DATE_NUMBER_FORMAT

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_instructions(sheet, categories: list[Category]) -> None:
    sheet.append(["Column", "What it means", "Data type / rules"])
    for name, meaning, rules in _COLUMN_INSTRUCTIONS:
        sheet.append([name, meaning, rules])

    sheet.append([])
    sheet.append(["Type", "Category"])
    for transaction_type, category_name in _type_category_rows(categories):
        sheet.append([transaction_type, category_name])


def _type_category_rows(categories: list[Category]) -> list[tuple[str, str]]:
    """(Type, Category) pairs for the Instructions sheet's table - every
    non-locked Category, Type-then-name ordered, built from the same live,
    locked-excluding list the Type/Category dropdowns use above so the table
    can never drift from what's actually selectable.
    """
    assignable = assignable_categories_by_type(categories)
    return [
        (transaction_type, category_name)
        for transaction_type in TYPE_ORDER
        for category_name in sorted(assignable.get(transaction_type, set()))
    ]


def from_payload(payload) -> Candidate:
    """The Candidate a request body describes.

    Raises ValueError - with a message naming what's wrong - for anything the
    caller could fix by sending a different body. Candidate's own
    __post_init__ raises the same way for a zero Amount or an invalid (Type,
    Category) pair, so the caller has one kind of error to handle.
    """
    if not isinstance(payload, dict):
        raise ValueError("Expected a JSON object describing one transaction")

    missing = [field for field in FIELDS if field not in payload]
    if missing:
        raise ValueError(f"Missing required field(s): {', '.join(missing)}")

    return Candidate(
        date=_date(payload["date"], "date"),
        amount=_number(payload["amount"], "amount"),
        type=payload["type"],
        category=payload["category"],
        notes=payload["notes"],
    )


def _number(value, field: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f"Field {field!r} must be a number, got {value!r}") from None


def _date(value, field: str) -> date:
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError):
        raise ValueError(f"Field {field!r} must be a date as YYYY-MM-DD, got {value!r}") from None


def candidate_payload(candidate: Candidate) -> dict:
    """The wire shape of one Candidate - the same fields `from_payload`
    reads back, with no `id` (a Candidate isn't stored yet). Round-trips the
    Import preview's to-write list to the commit endpoint (Issue #98).
    """
    return {
        "date": candidate.date.isoformat(),
        "amount": candidate.amount,
        "type": candidate.type,
        "category": candidate.category,
        "notes": candidate.notes,
    }


def decode_import_file(payload) -> bytes:
    """The raw `.xlsx` bytes `payload["file"]` (base64) decodes to.

    Raises ValueError - a single top-level error, per Issue #98's AC - for
    anything that isn't a JSON object carrying a base64 `file` string.
    """
    if not isinstance(payload, dict) or not isinstance(payload.get("file"), str):
        raise ValueError("Expected a JSON object with a base64-encoded 'file'")

    try:
        return base64.b64decode(payload["file"], validate=True)
    except binascii.Error:
        raise ValueError("Field 'file' must be valid base64") from None


def candidates_from_import_payload(payload) -> list[Candidate]:
    """The to-write Candidate list an Import-commit request body describes -
    exactly what the preview response's `candidates` handed back to the
    caller (Issue #98). Reuses `from_payload` per row, so a malformed entry
    fails the same way a malformed Add-transaction body would.
    """
    if not isinstance(payload, dict) or not isinstance(payload.get("candidates"), list):
        raise ValueError("Expected a JSON object with a 'candidates' array")

    return [from_payload(item) for item in payload["candidates"]]


def preview_import(file_bytes: bytes, categories: list[Category], existing_rows: list[ExistingRow]) -> dict:
    """Parse an uploaded Import workbook and classify every data row without
    writing anything (Issue #98) - `{"rows": [...], "candidates": [...]}`,
    where `rows` reports every data row's outcome (write / duplicate /
    rejected, with a reason for rejected) and `candidates` is the to-write
    list the commit endpoint expects back verbatim.

    Raises ValueError - a single top-level error - for a structurally
    invalid upload (missing/renamed columns, or bytes that aren't a real
    `.xlsx`); a row that's merely bad (invalid pair, unparsable date, ...) is
    reported per-row instead, never aborts the whole request.

    (Type, Category) validity and the locked-Category check are both against
    `categories` - the live table, not the fixed CATEGORIES_BY_TYPE dict - so
    a Category added through Category Management (Issue #91) is importable
    immediately, matching the Import template's own live dropdown (Issue
    #97) and how the store validates a manually-added transaction.
    """
    sheet = _load_transactions_sheet(file_bytes)
    type_by_category = type_lookup(categories)
    locked_categories = {c.name for c in categories if c.locked}

    parsed: list[tuple[int, Candidate]] = []
    rows: list[dict] = []
    for row_number, values in _data_rows(sheet):
        try:
            candidate = _candidate_from_row(values, type_by_category, locked_categories)
        except ValueError as cause:
            rows.append({"row": row_number, "outcome": "rejected", "reason": str(cause)})
            continue
        parsed.append((row_number, candidate))

    result = resolve_writes([candidate for _, candidate in parsed], existing_rows)
    to_write_ids = {id(candidate) for candidate in result.to_write}
    for row_number, candidate in parsed:
        outcome = "write" if id(candidate) in to_write_ids else "duplicate"
        rows.append({"row": row_number, "outcome": outcome})

    rows.sort(key=lambda row: row["row"])

    return {
        "rows": rows,
        "candidates": [candidate_payload(candidate) for candidate in result.to_write],
    }


def _load_transactions_sheet(file_bytes: bytes):
    """The Import template's "Transactions" data sheet - falling back to
    whichever sheet was active on save, in case a user renames it. Raises
    ValueError if `file_bytes` isn't a readable `.xlsx` at all, or if its
    header row doesn't match the template's exactly (a renamed/missing
    column - Issue #98's structural-error AC).
    """
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except (zipfile.BadZipFile, KeyError, OSError) as cause:
        raise ValueError("The uploaded file isn't a valid .xlsx workbook") from cause

    sheet = workbook["Transactions"] if "Transactions" in workbook.sheetnames else workbook.active

    header = tuple(cell.value for cell in sheet[1])
    if header != CSV_HEADER:
        seen = ", ".join(str(value) for value in header if value is not None) or "nothing"
        raise ValueError(f"Expected columns {', '.join(CSV_HEADER)} - got {seen}")

    return sheet


def _data_rows(sheet):
    """(row_number, [date, amount, type, category, notes]) for every data
    row that isn't entirely blank - the Import template pre-formats 500 rows
    (Issue #97), most of which a real upload leaves untouched.
    """
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row_number, column=column).value for column in range(1, 6)]
        if all(value is None or (isinstance(value, str) and value.strip() == "") for value in values):
            continue
        yield row_number, values


def _candidate_from_row(values: list, type_by_category: dict[str, str], locked_categories: set[str]) -> Candidate:
    date_value, amount_value, type_value, category_value, notes_value = values

    row_date = _import_date(date_value)
    amount = _import_amount(amount_value)
    transaction_type = str(type_value).strip() if type_value is not None else ""
    category = str(category_value).strip() if category_value is not None else ""
    notes = str(notes_value).strip() if notes_value is not None else ""

    if not transaction_type:
        raise ValueError("Type is required")
    if not category:
        raise ValueError("Category is required")
    if category in locked_categories:
        raise ValueError(f"Category {category!r} is locked and cannot be imported")
    if type_by_category.get(category) != transaction_type:
        raise ValueError(f"Category {category!r} is not a valid {transaction_type} Category")

    return Candidate(date=row_date, amount=amount, type=transaction_type, category=category, notes=notes)


def _import_date(value) -> date:
    """A genuine Excel date cell is accepted unambiguously; a text cell only
    as strict YYYY-MM-DD - anything else (an ambiguous DD/MM/YYYY-style
    string, a number, nothing) is rejected rather than guessed at (Issue
    #98's AC).
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            pass

    raise ValueError(f"Date must be a real Excel date or text as YYYY-MM-DD, got {value!r}")


def _import_amount(value) -> float:
    """A positive, non-zero number - matching the manual Add-transaction
    form's min="0.01" restriction (Issue #98's AC), stricter than Candidate's
    own __post_init__ (which only rejects zero, since a Statement Export
    Candidate may still carry a negative sign - see transaction_log.writer).
    """
    if isinstance(value, bool) or not isinstance(value, (int, float)) or value <= 0:
        raise ValueError(f"Amount must be a positive number, got {value!r}")
    return float(value)
