"""The Transactions tab's read and CRUD endpoints (Issues #33 and #35).

Exercised over real HTTP against a real LocalStore, mirroring
tests/dashboard/test_recurring_api.py - a test failing here means the
Dashboard's Transactions tab would fail the same way.
"""

import base64
import io
import json
from datetime import date
from urllib.error import HTTPError
from urllib.request import Request, urlopen

import openpyxl
import pytest

from transaction_log.categories import TYPE_ORDER, assignable_categories_by_type
from transaction_log.entries import Candidate

PAYLOAD = {
    "date": "2026-08-05",
    "amount": 42.5,
    "type": "Expense",
    "category": "Groceries",
    "notes": "Woolworths",
}


@pytest.fixture
def running_server(serve, store):
    return store, serve(store)


def get(server, path: str):
    with urlopen(f"http://127.0.0.1:{server.server_port}{path}") as response:
        return response.status, json.loads(response.read())


def call(server, method: str, path: str, payload: dict | None = None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    request = Request(
        f"http://127.0.0.1:{server.server_port}{path}",
        data=body,
        method=method,
        headers={"Content-Type": "application/json"} if body else {},
    )
    with urlopen(request) as response:
        raw = response.read()
        return response.status, (json.loads(raw) if raw else None)


def test_listing_on_an_empty_financial_year_returns_an_empty_list(running_server):
    _store, server = running_server

    status, body = get(server, "/api/transactions?year=2026&month=7")

    assert status == 200
    assert body == []


def test_listing_returns_the_financial_years_transactions_newest_first_with_ids(
    running_server, make_candidate
):
    store, server = running_server
    store.append_rows(
        [
            make_candidate(date=date(2026, 8, 1), amount=42.5, type="Expense", category="Groceries", notes="Woolworths"),
            make_candidate(date=date(2027, 3, 1), amount=4000.0, type="Income", category="Salary", notes="Employer"),
        ]
    )

    status, body = get(server, "/api/transactions?year=2026&month=7")

    assert status == 200
    assert [row["notes"] for row in body] == ["Employer", "Woolworths"]
    assert all(row["id"] is not None for row in body)
    assert body[0] == {
        "id": body[0]["id"],
        "date": "2027-03-01",
        "amount": 4000.0,
        "type": "Income",
        "category": "Salary",
        "notes": "Employer",
    }


def test_listing_excludes_transactions_outside_the_financial_year(running_server, make_candidate):
    store, server = running_server
    store.append_rows(
        [
            make_candidate(date=date(2026, 6, 30), notes="Last day of the prior Financial Year"),
            make_candidate(date=date(2027, 7, 1), notes="First day of the next Financial Year"),
            make_candidate(date=date(2026, 8, 15), notes="In the Financial Year"),
        ]
    )

    _status, body = get(server, "/api/transactions?year=2026&month=7")

    assert [row["notes"] for row in body] == ["In the Financial Year"]


def test_missing_query_params_returns_400(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        urlopen(f"http://127.0.0.1:{server.server_port}/api/transactions")

    assert exc_info.value.code == 400


def test_creating_a_transaction_returns_it_with_the_id_the_store_gave_it(running_server):
    _store, server = running_server

    status, created = call(server, "POST", "/api/transactions", PAYLOAD)

    assert status == 201
    assert created["id"] is not None
    assert {key: created[key] for key in PAYLOAD} == PAYLOAD


def test_a_created_transaction_then_appears_in_the_financial_year_listing(running_server):
    _store, server = running_server
    _status, created = call(server, "POST", "/api/transactions", PAYLOAD)

    _status, listed = call(server, "GET", "/api/transactions?year=2026&month=7")

    assert listed == [created]


def test_a_transaction_created_through_the_endpoint_is_what_dedupe_reads_next(running_server):
    store, server = running_server

    call(server, "POST", "/api/transactions", PAYLOAD)

    assert [row.notes for row in store.read_existing_rows()] == ["Woolworths"]


def test_editing_a_transaction_replaces_it_in_place(running_server):
    store, server = running_server
    _status, created = call(server, "POST", "/api/transactions", PAYLOAD)

    status, updated = call(server, "PUT", f"/api/transactions/{created['id']}", {**PAYLOAD, "amount": 50.0})

    assert status == 200
    assert updated["id"] == created["id"]
    assert updated["amount"] == 50.0
    assert [t.amount for t in store.read_transactions()] == [50.0]


def test_deleting_a_transaction_removes_it_from_the_listing(running_server):
    store, server = running_server
    _status, created = call(server, "POST", "/api/transactions", PAYLOAD)

    status, _body = call(server, "DELETE", f"/api/transactions/{created['id']}")

    assert status == 204
    assert call(server, "GET", "/api/transactions?year=2026&month=7")[1] == []
    assert store.read_transactions() == []


def test_creating_a_transaction_with_an_invalid_type_category_pair_is_rejected(running_server):
    store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        call(server, "POST", "/api/transactions", {**PAYLOAD, "category": "Salary"})

    assert exc_info.value.code == 400
    assert "Salary" in json.loads(exc_info.value.read())["error"]
    assert store.read_transactions() == []


def test_creating_a_transaction_with_a_zero_amount_is_rejected(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        call(server, "POST", "/api/transactions", {**PAYLOAD, "amount": 0})

    assert exc_info.value.code == 400


def test_editing_a_transaction_into_an_invalid_pair_is_rejected_and_leaves_it_alone(running_server):
    store, server = running_server
    _status, created = call(server, "POST", "/api/transactions", PAYLOAD)

    with pytest.raises(HTTPError) as exc_info:
        call(server, "PUT", f"/api/transactions/{created['id']}", {**PAYLOAD, "category": "Salary"})

    assert exc_info.value.code == 400
    assert [t.category for t in store.read_transactions()] == ["Groceries"]


def test_a_payload_missing_a_field_is_rejected_with_a_clear_error(running_server):
    _store, server = running_server
    incomplete = {key: value for key, value in PAYLOAD.items() if key != "date"}

    with pytest.raises(HTTPError) as exc_info:
        call(server, "POST", "/api/transactions", incomplete)

    assert exc_info.value.code == 400
    assert "date" in json.loads(exc_info.value.read())["error"]


def test_a_body_that_is_not_json_is_rejected(running_server):
    _store, server = running_server
    request = Request(
        f"http://127.0.0.1:{server.server_port}/api/transactions",
        data=b"not json at all",
        method="POST",
    )

    with pytest.raises(HTTPError) as exc_info:
        urlopen(request)

    assert exc_info.value.code == 400


def test_editing_a_transaction_that_does_not_exist_returns_404(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        call(server, "PUT", "/api/transactions/404", PAYLOAD)

    assert exc_info.value.code == 404


def test_deleting_a_transaction_that_does_not_exist_returns_404(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        call(server, "DELETE", "/api/transactions/404")

    assert exc_info.value.code == 404


def test_a_non_numeric_transaction_id_returns_404(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        call(server, "DELETE", "/api/transactions/not-a-number")

    assert exc_info.value.code == 404


def get_csv(server, path: str):
    with urlopen(f"http://127.0.0.1:{server.server_port}{path}") as response:
        return response.status, response.headers, response.read().decode("utf-8")


def test_export_on_an_empty_range_returns_a_header_only_csv(running_server):
    _store, server = running_server

    status, _headers, body = get_csv(server, "/api/transactions/export?start=2026-08-01&end=2026-08-31")

    assert status == 200
    assert body == "Date,Amount,Type,Category,Notes\r\n"


def test_export_has_exactly_five_columns_and_no_id(running_server):
    _store, server = running_server
    call(server, "POST", "/api/transactions", PAYLOAD)

    _status, _headers, body = get_csv(server, "/api/transactions/export?start=2026-08-01&end=2026-08-31")

    lines = body.strip("\r\n").split("\r\n")
    assert lines[0] == "Date,Amount,Type,Category,Notes"
    assert lines[1] == "2026-08-05,42.5,Expense,Groceries,Woolworths"


def test_export_bounds_are_inclusive(running_server, make_candidate):
    store, server = running_server
    store.append_rows(
        [
            make_candidate(date=date(2026, 7, 31), notes="Before the range"),
            make_candidate(date=date(2026, 8, 1), notes="Start of the range"),
            make_candidate(date=date(2026, 8, 31), notes="End of the range"),
            make_candidate(date=date(2026, 9, 1), notes="After the range"),
        ]
    )

    _status, _headers, body = get_csv(server, "/api/transactions/export?start=2026-08-01&end=2026-08-31")

    assert "Before the range" not in body
    assert "Start of the range" in body
    assert "End of the range" in body
    assert "After the range" not in body


def test_export_spans_a_financial_year_boundary(running_server, make_candidate):
    store, server = running_server
    store.append_rows(
        [
            make_candidate(date=date(2026, 6, 30), notes="Last day of FY25-26"),
            make_candidate(date=date(2026, 7, 1), notes="First day of FY26-27"),
        ]
    )

    _status, _headers, body = get_csv(server, "/api/transactions/export?start=2026-06-30&end=2026-07-01")

    assert "Last day of FY25-26" in body
    assert "First day of FY26-27" in body


def test_export_sets_content_disposition_attachment(running_server):
    _store, server = running_server

    _status, headers, _body = get_csv(server, "/api/transactions/export?start=2026-08-01&end=2026-08-31")

    assert headers["Content-Disposition"].startswith("attachment;")
    assert ".csv" in headers["Content-Disposition"]


def test_export_missing_query_params_returns_400(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        urlopen(f"http://127.0.0.1:{server.server_port}/api/transactions/export")

    assert exc_info.value.code == 400


def test_export_malformed_date_returns_400(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        urlopen(
            f"http://127.0.0.1:{server.server_port}/api/transactions/export?start=not-a-date&end=2026-08-31"
        )

    assert exc_info.value.code == 400


def get_workbook(server, path: str):
    """(status, headers, openpyxl Workbook) for a GET against `path`."""
    with urlopen(f"http://127.0.0.1:{server.server_port}{path}") as response:
        status, headers, body = response.status, response.headers, response.read()
    return status, headers, openpyxl.load_workbook(io.BytesIO(body))


def category_dropdown_values(workbook) -> list[str]:
    lists_sheet = workbook["Lists"]
    return [
        lists_sheet.cell(row=row, column=2).value
        for row in range(1, lists_sheet.max_row + 1)
        if lists_sheet.cell(row=row, column=2).value is not None
    ]


def instructions_type_category_rows(workbook) -> list[tuple[str, str]]:
    rows = list(workbook["Instructions"].iter_rows(values_only=True))
    header_index = rows.index(("Type", "Category", None))
    return [(row[0], row[1]) for row in rows[header_index + 1 :] if row[0] is not None]


IMPORT_TEMPLATE_PATH = "/api/transactions/import-template"


def test_import_template_sets_content_disposition_attachment(running_server):
    _store, server = running_server

    with urlopen(f"http://127.0.0.1:{server.server_port}{IMPORT_TEMPLATE_PATH}") as response:
        status, headers = response.status, response.headers

    assert status == 200
    assert headers["Content-Disposition"].startswith("attachment;")
    assert headers["Content-Disposition"].endswith('.xlsx"')


def test_import_template_data_sheet_has_the_transaction_log_columns_in_order(running_server):
    _store, server = running_server

    _status, _headers, workbook = get_workbook(server, IMPORT_TEMPLATE_PATH)

    sheet = workbook["Transactions"]
    assert [cell.value for cell in sheet[1]] == ["Date", "Amount", "Type", "Category", "Notes"]


def test_import_template_date_column_is_formatted_as_a_real_date_column(running_server):
    _store, server = running_server

    _status, _headers, workbook = get_workbook(server, IMPORT_TEMPLATE_PATH)

    sheet = workbook["Transactions"]
    assert sheet.cell(row=2, column=1).number_format == "YYYY-MM-DD"
    assert sheet.cell(row=50, column=1).number_format == "YYYY-MM-DD"


def test_import_template_type_dropdown_offers_the_four_fixed_types(running_server):
    _store, server = running_server

    _status, _headers, workbook = get_workbook(server, IMPORT_TEMPLATE_PATH)

    lists_sheet = workbook["Lists"]
    assert lists_sheet.sheet_state == "hidden"
    type_values = [lists_sheet.cell(row=row, column=1).value for row in range(1, 5)]
    assert type_values == list(TYPE_ORDER)


def test_import_template_category_dropdown_matches_live_categories_and_excludes_locked(running_server):
    store, server = running_server
    store.create_category("Expense", "Pet Care", None)

    _status, _headers, workbook = get_workbook(server, IMPORT_TEMPLATE_PATH)

    expected = sorted({c.name for c in store.read_categories() if not c.locked})
    assert category_dropdown_values(workbook) == expected
    assert "Beem Adjustment" not in expected  # sanity: the live fixture really is locked
    assert "Beem Adjustment" not in category_dropdown_values(workbook)


def test_import_template_instructions_type_category_table_matches_live_categories(running_server):
    store, server = running_server

    _status, _headers, workbook = get_workbook(server, IMPORT_TEMPLATE_PATH)

    assignable = assignable_categories_by_type(store.read_categories())
    expected = [(t, name) for t in TYPE_ORDER for name in sorted(assignable.get(t, set()))]

    rows = instructions_type_category_rows(workbook)
    assert rows == expected
    assert all(name != "Beem Adjustment" for _t, name in rows)


def test_import_template_reflects_a_newly_added_category_with_no_restart(running_server):
    store, server = running_server
    _status, _headers, before = get_workbook(server, IMPORT_TEMPLATE_PATH)
    assert "Pet Care" not in category_dropdown_values(before)
    assert not any(name == "Pet Care" for _t, name in instructions_type_category_rows(before))

    store.create_category("Expense", "Pet Care", None)

    _status, _headers, after = get_workbook(server, IMPORT_TEMPLATE_PATH)
    assert "Pet Care" in category_dropdown_values(after)
    assert ("Expense", "Pet Care") in instructions_type_category_rows(after)


# Issue #98 - upload/preview/confirm.

IMPORT_PREVIEW_PATH = "/api/transactions/import-preview"
IMPORT_COMMIT_PATH = "/api/transactions/import-commit"


def build_import_workbook(rows: list[tuple]) -> bytes:
    """An in-memory `.xlsx` shaped like the Import template - same header,
    same "Transactions" sheet name - with `rows` written as data rows
    starting at row 2. Each row is a (date, amount, type, category, notes)
    tuple; `date` may be a `date`/`datetime` (a genuine Excel date cell) or a
    plain string (a text cell).
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(["Date", "Amount", "Type", "Category", "Notes"])
    for row in rows:
        sheet.append(list(row))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def upload_payload(file_bytes: bytes) -> dict:
    return {"file": base64.b64encode(file_bytes).decode("ascii")}


def outcomes_by_row(body: dict) -> dict[int, dict]:
    return {row["row"]: row for row in body["rows"]}


def test_import_preview_classifies_a_mixed_files_rows_correctly(running_server):
    store, server = running_server
    store.append_rows(
        [
            _candidate(date=date(2026, 8, 1), amount=15.0, type="Expense", category="Groceries", notes="Existing row")
        ]
    )

    file_bytes = build_import_workbook(
        [
            (date(2026, 8, 10), 55.0, "Expense", "Groceries", "New row"),  # row 2: write
            (date(2026, 8, 1), 15.0, "Expense", "Groceries", "Existing row"),  # row 3: duplicate
            (date(2026, 8, 11), 20.0, "Expense", "Salary", "Bad pair"),  # row 4: rejected - bad pair
            (date(2026, 8, 12), 0, "Expense", "Groceries", "Zero amount"),  # row 5: rejected - zero
            (date(2026, 8, 13), -5.0, "Expense", "Groceries", "Negative amount"),  # row 6: rejected - negative
            (date(2026, 8, 14), 30.0, "Expense", "Beem Adjustment", "Locked category"),  # row 7: rejected - locked
            ("11/08/2026", 40.0, "Expense", "Groceries", "Ambiguous date"),  # row 8: rejected - ambiguous date
            (date(2026, 8, 15), 60.0, "Expense", "Groceries", "Real Excel date"),  # row 9: write
        ]
    )

    status, body = call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))

    assert status == 200
    outcomes = outcomes_by_row(body)
    assert outcomes[2]["outcome"] == "write"
    assert outcomes[3]["outcome"] == "duplicate"
    assert outcomes[4]["outcome"] == "rejected"
    assert "Salary" in outcomes[4]["reason"]
    assert outcomes[5]["outcome"] == "rejected"
    assert outcomes[6]["outcome"] == "rejected"
    assert outcomes[7]["outcome"] == "rejected"
    assert "Beem Adjustment" in outcomes[7]["reason"]
    assert outcomes[8]["outcome"] == "rejected"
    assert outcomes[9]["outcome"] == "write"

    assert [c["notes"] for c in body["candidates"]] == ["New row", "Real Excel date"]

    # A preview writes nothing - only the row seeded before the request exists.
    assert [t.notes for t in store.read_transactions()] == ["Existing row"]


def test_import_preview_writes_nothing(running_server):
    store, server = running_server
    file_bytes = build_import_workbook([(date(2026, 8, 10), 55.0, "Expense", "Groceries", "New row")])

    call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))

    assert store.read_transactions() == []


def test_import_preview_skips_fully_blank_rows(running_server):
    _store, server = running_server
    file_bytes = build_import_workbook(
        [
            (date(2026, 8, 10), 55.0, "Expense", "Groceries", "New row"),
            (None, None, None, None, None),
        ]
    )

    status, body = call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))

    assert status == 200
    assert [row["row"] for row in body["rows"]] == [2]


def test_import_preview_rejects_a_structurally_malformed_upload_with_one_top_level_error(running_server):
    _store, server = running_server

    with pytest.raises(HTTPError) as exc_info:
        call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(b"not an xlsx file at all"))

    assert exc_info.value.code == 400
    body = json.loads(exc_info.value.read())
    assert "error" in body
    assert "rows" not in body


def test_import_preview_rejects_a_workbook_with_renamed_columns(running_server):
    _store, server = running_server
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(["Date", "Amount", "Kind", "Category", "Notes"])
    sheet.append([date(2026, 8, 10), 55.0, "Expense", "Groceries", "New row"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(HTTPError) as exc_info:
        call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(buffer.getvalue()))

    assert exc_info.value.code == 400
    assert "error" in json.loads(exc_info.value.read())


def test_import_commit_writes_exactly_the_previewed_to_write_candidates(running_server):
    store, server = running_server
    file_bytes = build_import_workbook(
        [
            (date(2026, 8, 10), 55.0, "Expense", "Groceries", "New row"),
            (date(2026, 8, 11), 20.0, "Income", "Salary", "Payday"),
        ]
    )
    _status, preview = call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))

    status, body = call(server, "POST", IMPORT_COMMIT_PATH, {"candidates": preview["candidates"]})

    assert status == 200
    assert {t.notes for t in store.read_transactions()} == {"New row", "Payday"}
    assert {row["notes"] for row in body["written"]} == {"New row", "Payday"}


def test_import_commit_reskips_a_row_that_became_a_duplicate_since_preview(running_server):
    store, server = running_server
    file_bytes = build_import_workbook([(date(2026, 8, 10), 55.0, "Expense", "Groceries", "New row")])
    _status, preview = call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))

    # The row is written by some other means between preview and commit.
    store.append_rows(
        [_candidate(date=date(2026, 8, 10), amount=55.0, type="Expense", category="Groceries", notes="New row")]
    )

    status, body = call(server, "POST", IMPORT_COMMIT_PATH, {"candidates": preview["candidates"]})

    assert status == 200
    assert body["written"] == []
    assert len(store.read_transactions()) == 1


def test_reimporting_a_file_already_present_reports_every_row_as_duplicate_and_writes_nothing(running_server):
    store, server = running_server
    file_bytes = build_import_workbook([(date(2026, 8, 10), 55.0, "Expense", "Groceries", "New row")])
    call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))
    _status, first_preview = call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))
    call(server, "POST", IMPORT_COMMIT_PATH, {"candidates": first_preview["candidates"]})

    status, second_preview = call(server, "POST", IMPORT_PREVIEW_PATH, upload_payload(file_bytes))

    assert status == 200
    assert [row["outcome"] for row in second_preview["rows"]] == ["duplicate"]
    assert second_preview["candidates"] == []
    assert len(store.read_transactions()) == 1


def _candidate(**kwargs):
    return Candidate(**kwargs)
